#!/usr/bin/env python3
"""Build an interactive Excel dashboard from raw consultant data quality export."""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

METRIC_TYPES = ["Weighted", "Absolute"]
SEGMENT_DIMENSIONS = ["Department", "Team", "Unit", "Legal entity", "Location", "Lead"]

REQUIRED_HEADERS = [
    "External",
    "Mat",
    "Url",
    "Full name",
    "Email",
    "Entry date",
    "Work experience since",
    "Total engagements",
    "Engagements with invalid dates",
    "Ongoing engagements",
    "Oldest ongoing engagement start date",
    "Last finished engagement start date",
    "Last finished engagement end date",
    "Is profile photo missing",
    "Months since entry baseline",
    "Absolute months since entry",
    "Weighted months since entry",
    "Absolute missing months since entry",
    "Weighted missing months since entry",
    "Absolute coverage ratio since entry",
    "Weighted coverage ratio since entry",
    "Months before entry baseline",
    "Absolute months before entry",
    "Weighted months before entry",
    "Absolute missing months before entry",
    "Weighted missing months before entry",
    "Absolute coverage ratio before entry",
    "Weighted coverage ratio before entry",
    "Is entry date missing",
    "Is work experience since missing",
    "Is work experience since after entry date",
    "Is available",
    "Available from",
    "Available to",
    "Is willing to travel",
    "Available days per week",
    "Availability comment",
    "Department",
    "Team",
    "Unit",
    "Legal entity",
    "Location",
    "Lead",
]



def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build Excel dashboard from raw data quality export")
    parser.add_argument("--input", required=True, help="Path to input workbook")
    parser.add_argument("--output", required=True, help="Path to output workbook")
    return parser.parse_args()



def load_source(input_path: Path) -> Tuple[List[str], List[List[object]]]:
    source_wb = load_workbook(input_path, data_only=False)
    candidate_sheets = [source_wb.active]
    if "Raw_data" in source_wb.sheetnames and source_wb["Raw_data"] not in candidate_sheets:
        candidate_sheets.append(source_wb["Raw_data"])

    source_ws = None
    headers: List[str] = []
    missing: List[str] = REQUIRED_HEADERS.copy()

    for sheet in candidate_sheets:
        sheet_headers = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
        sheet_headers = [str(h).strip() if h is not None else "" for h in sheet_headers]
        sheet_missing = [h for h in REQUIRED_HEADERS if h not in sheet_headers]
        if not sheet_missing:
            source_ws = sheet
            headers = sheet_headers
            missing = []
            break

    if source_ws is None:
        raise ValueError(f"Input workbook missing required columns: {missing}")

    rows: List[List[object]] = []
    for row in source_ws.iter_rows(min_row=2, max_row=source_ws.max_row, min_col=1, max_col=len(headers), values_only=True):
        if any(v is not None and v != "" for v in row):
            rows.append(list(row))

    return headers, rows



def set_column_widths(ws, widths: Dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width



def write_raw_data_sheet(wb: Workbook, headers: Sequence[str], rows: Sequence[Sequence[object]]) -> Tuple[Dict[str, str], int]:
    ws = wb.active
    ws.title = "Raw_data"

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, row_values in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    max_row = len(rows) + 1
    max_col = len(headers)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(header) + 2, 12), 38)

    header_to_col = {h: get_column_letter(i + 1) for i, h in enumerate(headers)}
    return header_to_col, max_row



def write_lists_sheet(wb: Workbook, rows: Sequence[Sequence[object]], headers: Sequence[str]) -> Dict[str, int]:
    ws = wb.create_sheet("Lists")
    ws.sheet_state = "hidden"

    idx = {h: i for i, h in enumerate(headers)}
    list_defs = [
        ("Metric_type", METRIC_TYPES, "A"),
        ("Department", None, "B"),
        ("Team", None, "C"),
        ("Unit", None, "D"),
        ("Legal_entity", None, "E"),
        ("Location", None, "F"),
        ("Lead", None, "G"),
        ("Is_available", None, "H"),
        ("Segment_dimension", SEGMENT_DIMENSIONS, "I"),
    ]

    end_rows: Dict[str, int] = {}

    for list_name, static_values, col in list_defs:
        ws[f"{col}1"] = list_name

        if static_values is not None:
            values = list(static_values)
        else:
            source_header = list_name.replace("_", " ")
            source_idx = idx[source_header]
            unique = sorted(
                {
                    str(row[source_idx]).strip()
                    for row in rows
                    if row[source_idx] is not None and str(row[source_idx]).strip() != ""
                }
            )
            values = ["All"] + unique

        for r, value in enumerate(values, start=2):
            ws[f"{col}{r}"] = value

        end_rows[list_name] = len(values) + 1

    set_column_widths(
        ws,
        {
            "A": 16,
            "B": 44,
            "C": 44,
            "D": 44,
            "E": 44,
            "F": 44,
            "G": 32,
            "H": 14,
            "I": 20,
        },
    )

    return end_rows



def write_model_sheet(wb: Workbook, max_raw_row: int, raw_cols: Dict[str, str], list_end_rows: Dict[str, int]) -> None:
    ws = wb.create_sheet("Model")
    ws.sheet_state = "hidden"

    ws.append(
        [
            "Include",
            "Selected_ratio_since",
            "Selected_ratio_before",
            "Ongoing_age_months",
            "Availability_inconsistency",
            "Anomaly_flag",
            "Segment_value",
            "Low_coverage_flag",
            "Anomaly_score",
            "Candidate_segment",
            "Candidate_total",
            "Candidate_low",
            "Candidate_low_rate",
            "Candidate_rank_key",
            "Top_segment",
            "Top_low_rate",
            "Top_count",
            "Rank_key",
            "Selected_since_included",
            "Selected_before_included",
            "Weighted_since_included",
            "Weighted_before_included",
        ]
    )

    start_row = 2
    end_row = max_raw_row

    dep_list = f"Lists!$B$3:$B${max(3, list_end_rows['Department'])}"
    team_list = f"Lists!$C$3:$C${max(3, list_end_rows['Team'])}"
    unit_list = f"Lists!$D$3:$D${max(3, list_end_rows['Unit'])}"
    legal_list = f"Lists!$E$3:$E${max(3, list_end_rows['Legal_entity'])}"
    location_list = f"Lists!$F$3:$F${max(3, list_end_rows['Location'])}"
    lead_list = f"Lists!$G$3:$G${max(3, list_end_rows['Lead'])}"

    for r in range(start_row, end_row + 1):
        ws[f"A{r}"] = (
            f"=--(AND("
            f"OR(Dashboard!$B$4=\"All\",Raw_data!${raw_cols['Department']}{r}=Dashboard!$B$4),"
            f"OR(Dashboard!$B$5=\"All\",Raw_data!${raw_cols['Team']}{r}=Dashboard!$B$5),"
            f"OR(Dashboard!$B$6=\"All\",Raw_data!${raw_cols['Unit']}{r}=Dashboard!$B$6),"
            f"OR(Dashboard!$B$7=\"All\",Raw_data!${raw_cols['Legal entity']}{r}=Dashboard!$B$7),"
            f"OR(Dashboard!$B$8=\"All\",Raw_data!${raw_cols['Location']}{r}=Dashboard!$B$8),"
            f"OR(Dashboard!$B$9=\"All\",Raw_data!${raw_cols['Lead']}{r}=Dashboard!$B$9),"
            f"OR(Dashboard!$B$10=\"All\",Raw_data!${raw_cols['Is available']}{r}=Dashboard!$B$10),"
            f"OR(Raw_data!${raw_cols['Entry date']}{r}=\"\",Raw_data!${raw_cols['Entry date']}{r}<=TODAY())"
            f"))"
        )

        ws[f"B{r}"] = (
            f"=IF(Dashboard!$B$3=\"Weighted\",Raw_data!${raw_cols['Weighted coverage ratio since entry']}{r},"
            f"Raw_data!${raw_cols['Absolute coverage ratio since entry']}{r})"
        )
        ws[f"C{r}"] = (
            f"=IF(Dashboard!$B$3=\"Weighted\",Raw_data!${raw_cols['Weighted coverage ratio before entry']}{r},"
            f"Raw_data!${raw_cols['Absolute coverage ratio before entry']}{r})"
        )

        ws[f"D{r}"] = (
            f"=IF(Raw_data!${raw_cols['Oldest ongoing engagement start date']}{r}=\"\",\"\","
            f"IF(Raw_data!${raw_cols['Oldest ongoing engagement start date']}{r}>TODAY(),0,"
            f"DATEDIF(Raw_data!${raw_cols['Oldest ongoing engagement start date']}{r},TODAY(),\"M\")))"
        )

        ws[f"E{r}"] = (
            f"=--(AND(Raw_data!${raw_cols['Is available']}{r}=\"No\","
            f"OR(Raw_data!${raw_cols['Available from']}{r}<>\"\","
            f"Raw_data!${raw_cols['Available days per week']}{r}>0,"
            f"Raw_data!${raw_cols['Availability comment']}{r}<>\"\")))"
        )

        ws[f"F{r}"] = (
            f"=--(OR(AND(ISNUMBER(B{r}),B{r}<0.5,OR(D{r}>=24,Raw_data!${raw_cols['Engagements with invalid dates']}{r}>0)),"
            f"Raw_data!${raw_cols['Is work experience since after entry date']}{r}=TRUE))"
        )

        ws[f"G{r}"] = (
            f"=IF(Dashboard!$B$11=\"Department\",Raw_data!${raw_cols['Department']}{r},"
            f"IF(Dashboard!$B$11=\"Team\",Raw_data!${raw_cols['Team']}{r},"
            f"IF(Dashboard!$B$11=\"Unit\",Raw_data!${raw_cols['Unit']}{r},"
            f"IF(Dashboard!$B$11=\"Legal entity\",Raw_data!${raw_cols['Legal entity']}{r},"
            f"IF(Dashboard!$B$11=\"Location\",Raw_data!${raw_cols['Location']}{r},Raw_data!${raw_cols['Lead']}{r})))))"
        )

        ws[f"H{r}"] = f"=--(AND(ISNUMBER(B{r}),B{r}<0.5))"

        ws[f"I{r}"] = (
            f"=MAX(0,(0.5-B{r})*100)+"
            f"IF(D{r}>=24,20,0)+"
            f"IF(Raw_data!${raw_cols['Engagements with invalid dates']}{r}>0,20,0)+"
            f"IF(Raw_data!${raw_cols['Is work experience since after entry date']}{r}=TRUE,10,0)"
        )

        ws[f"R{r}"] = f"=IF(A{r}=1,I{r}+ROW()/100000000,\"\")"
        ws[f"S{r}"] = f"=IF(A{r}=1,Raw_data!${raw_cols['Absolute coverage ratio since entry']}{r},\"\")"
        ws[f"T{r}"] = f"=IF(A{r}=1,Raw_data!${raw_cols['Absolute coverage ratio before entry']}{r},\"\")"
        ws[f"U{r}"] = f"=IF(A{r}=1,Raw_data!${raw_cols['Weighted coverage ratio since entry']}{r},\"\")"
        ws[f"V{r}"] = f"=IF(A{r}=1,Raw_data!${raw_cols['Weighted coverage ratio before entry']}{r},\"\")"

    choose_expr = (
        "CHOOSE(MATCH(Dashboard!$B$11,{\"Department\",\"Team\",\"Unit\",\"Legal entity\",\"Location\",\"Lead\"},0),"
        f"INDEX({dep_list},ROW()-1),"
        f"INDEX({team_list},ROW()-1),"
        f"INDEX({unit_list},ROW()-1),"
        f"INDEX({legal_list},ROW()-1),"
        f"INDEX({location_list},ROW()-1),"
        f"INDEX({lead_list},ROW()-1))"
    )

    for r in range(2, 1002):
        ws[f"J{r}"] = f"=IFERROR({choose_expr},\"\")"
        ws[f"K{r}"] = f"=IF(J{r}=\"\",\"\",SUMPRODUCT(($A$2:$A${end_row}=1)*($G$2:$G${end_row}=J{r})))"
        ws[f"L{r}"] = f"=IF(J{r}=\"\",\"\",SUMPRODUCT(($A$2:$A${end_row}=1)*($G$2:$G${end_row}=J{r})*(ISNUMBER($B$2:$B${end_row}))*($B$2:$B${end_row}<0.5)))"
        ws[f"M{r}"] = f"=IFERROR(L{r}/K{r},\"\")"
        ws[f"N{r}"] = f"=IF(K{r}=\"\",\"\",M{r}+K{r}/100000+ROW()/100000000)"

    for r in range(2, 12):
        k = r - 1
        ws[f"O{r}"] = f"=IFERROR(INDEX($J$2:$J$1001,MATCH(LARGE($N$2:$N$1001,{k}),$N$2:$N$1001,0)),\"\")"
        ws[f"P{r}"] = f"=IFERROR(INDEX($M$2:$M$1001,MATCH(LARGE($N$2:$N$1001,{k}),$N$2:$N$1001,0)),\"\")"
        ws[f"Q{r}"] = f"=IFERROR(INDEX($K$2:$K$1001,MATCH(LARGE($N$2:$N$1001,{k}),$N$2:$N$1001,0)),\"\")"

    set_column_widths(
        ws,
        {
            "A": 10,
            "B": 18,
            "C": 18,
            "D": 16,
            "E": 24,
            "F": 12,
            "G": 24,
            "H": 14,
            "I": 14,
            "J": 24,
            "K": 14,
            "L": 12,
            "M": 14,
            "N": 16,
            "O": 24,
            "P": 14,
            "Q": 12,
            "R": 16,
            "S": 16,
            "T": 16,
            "U": 16,
            "V": 16,
        },
    )



def add_dropdown(ws, cell: str, formula: str) -> None:
    dv = DataValidation(type="list", formula1=formula, allow_blank=False)
    dv.error = "Select a value from the dropdown"
    dv.errorTitle = "Invalid selection"
    dv.prompt = "Choose a filter value"
    dv.promptTitle = "Filter"
    ws.add_data_validation(dv)
    dv.add(cell)



def add_comments(ws) -> None:
    filter_comments = {
        "A3": "Choose whether ratio-based charts use Weighted or Absolute metrics.",
        "A4": "Filter to one Department or keep All.",
        "A5": "Filter to one Team or keep All.",
        "A6": "Filter to one Unit or keep All.",
        "A7": "Filter to one Legal Entity or keep All.",
        "A8": "Filter to one Office Location or keep All.",
        "A9": "Filter by Lead owner or keep All.",
        "A10": "Filter availability bucket (Yes/No/-) or keep All.",
        "A11": "Choose the dimension used in the Top 10 segment chart.",
    }

    kpi_comments = {
        "D6": "Median weighted coverage ratio since entry date for currently filtered consultants.",
        "D7": "Median absolute coverage ratio since entry date for currently filtered consultants.",
        "D8": "Median weighted coverage ratio before entry date for currently filtered consultants.",
        "D9": "Median absolute coverage ratio before entry date for currently filtered consultants.",
        "H5": "Count of filtered consultants whose oldest ongoing engagement started at least 24 months ago.",
        "H6": "Count of filtered consultants with at least one invalid-date engagement.",
        "H7": "A consultant is counted as an anomaly when at least one of these applies: (1) Since-entry coverage is below 0.50 and there is either an ongoing engagement older than 24 months or at least one invalid-date engagement, or (2) Work experience since is later than entry date. This helps spot likely profile maintenance or timeline issues.",
    }

    band_comments = {
        "A14": "Consultants with since-entry coverage ratio below 0.30 (less than 30% covered).",
        "A15": "Consultants with since-entry coverage ratio from 0.30 up to 0.49.",
        "A16": "Consultants with since-entry coverage ratio from 0.50 up to 0.79.",
        "A17": "Consultants with since-entry coverage ratio at or above 0.80.",
        "D14": "Consultants with before-entry coverage ratio below 0.30 (less than 30% covered).",
        "D15": "Consultants with before-entry coverage ratio from 0.30 up to 0.49.",
        "D16": "Consultants with before-entry coverage ratio from 0.50 up to 0.79.",
        "D17": "Consultants with before-entry coverage ratio at or above 0.80.",
    }

    section_comments = {
        "A21": "Low coverage in this chart means coverage ratio below 0.50. That combines both low bands: <0.30 and 0.30-0.49.",
        "A35": "This list is sorted anomaly-first. The same anomaly logic is used as above: low since-entry coverage (<0.50) with stale ongoing and/or invalid dates, or work-experience-since after entry date.",
    }

    for cell, text in {**filter_comments, **kpi_comments, **band_comments, **section_comments}.items():
        comment = Comment(text, "Codex")
        # Keep most notes compact; only anomaly explanations should be larger.
        if cell in {"H7", "A35"}:
            comment.width = 760
            comment.height = 240
        elif len(text) <= 90:
            comment.width = 320
            comment.height = 90
        elif len(text) <= 180:
            comment.width = 420
            comment.height = 120
        else:
            comment.width = 560
            comment.height = 170
        ws[cell].comment = comment



def write_dashboard_sheet(wb: Workbook, max_raw_row: int, list_end_rows: Dict[str, int], raw_cols: Dict[str, str]) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Internal Consultant Data Quality Dashboard"
    ws["A1"].font = Font(size=18, bold=True, color="1F4E78")

    # Filter panel
    labels = [
        "Metric_type",
        "Department",
        "Team",
        "Unit",
        "Legal_entity",
        "Location",
        "Lead",
        "Is_available",
        "Segment_dimension",
    ]
    for row, label in enumerate(labels, start=3):
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, color="1F4E78")

    ws["B3"] = "Weighted"
    ws["B4"] = "All"
    ws["B5"] = "All"
    ws["B6"] = "All"
    ws["B7"] = "All"
    ws["B8"] = "All"
    ws["B9"] = "All"
    ws["B10"] = "All"
    ws["B11"] = "Department"

    add_dropdown(ws, "B3", "=Lists!$A$2:$A$3")
    add_dropdown(ws, "B4", f"=Lists!$B$2:$B${list_end_rows['Department']}")
    add_dropdown(ws, "B5", f"=Lists!$C$2:$C${list_end_rows['Team']}")
    add_dropdown(ws, "B6", f"=Lists!$D$2:$D${list_end_rows['Unit']}")
    add_dropdown(ws, "B7", f"=Lists!$E$2:$E${list_end_rows['Legal_entity']}")
    add_dropdown(ws, "B8", f"=Lists!$F$2:$F${list_end_rows['Location']}")
    add_dropdown(ws, "B9", f"=Lists!$G$2:$G${list_end_rows['Lead']}")
    add_dropdown(ws, "B10", f"=Lists!$H$2:$H${list_end_rows['Is_available']}")
    add_dropdown(ws, "B11", f"=Lists!$I$2:$I${list_end_rows['Segment_dimension']}")

    for cell in ["B3", "B4", "B5", "B6", "B7", "B8", "B9", "B10", "B11"]:
        ws[cell].fill = PatternFill(fill_type="solid", start_color="E8F1FA", end_color="E8F1FA")

    ws["D3"] = "Selected metric"
    ws["E3"] = "=B3"
    ws["D3"].font = Font(bold=True)
    ws["E3"].font = Font(bold=True, color="1F4E78")

    end_row = max_raw_row

    # KPI block
    ws["D5"] = "Consultants (filtered)"
    ws["E5"] = f"=SUM(Model!$A$2:$A${end_row})"

    ws["D6"] = "Weighted median since entry"
    ws["E6"] = f"=IFERROR(MEDIAN(Model!$U$2:$U${end_row}),\"\")"

    ws["D7"] = "Absolute median since entry"
    ws["E7"] = f"=IFERROR(MEDIAN(Model!$S$2:$S${end_row}),\"\")"

    ws["D8"] = "Weighted median before entry"
    ws["E8"] = f"=IFERROR(MEDIAN(Model!$V$2:$V${end_row}),\"\")"

    ws["D9"] = "Absolute median before entry"
    ws["E9"] = f"=IFERROR(MEDIAN(Model!$T$2:$T${end_row}),\"\")"

    ws["H5"] = "Ongoing >=24 months"
    ws["I5"] = f"=SUMPRODUCT((Model!$A$2:$A${end_row}=1)*(Model!$D$2:$D${end_row}>=24))"

    ws["H6"] = "Invalid-date consultants"
    ws["I6"] = f"=SUMPRODUCT((Model!$A$2:$A${end_row}=1)*(Raw_data!$I$2:$I${end_row}>0))"

    ws["H7"] = "Anomaly consultants"
    ws["I7"] = f"=SUMPRODUCT((Model!$A$2:$A${end_row}=1)*(Model!$F$2:$F${end_row}=1))"

    for row in range(5, 10):
        ws[f"D{row}"].font = Font(bold=True)
    for row in range(5, 8):
        ws[f"H{row}"].font = Font(bold=True)
    for c in ["E6", "E7", "E8", "E9"]:
        ws[c].number_format = "0.00"

    # Band tables
    ws["A13"] = "Coverage since entry bands"
    ws["A13"].font = Font(bold=True, color="1F4E78")
    ws["B13"] = "Consultants"

    ws["D13"] = "Coverage before entry bands"
    ws["D13"].font = Font(bold=True, color="1F4E78")
    ws["E13"] = "Consultants"

    ws["H13"] = "Ongoing age buckets"
    ws["H13"].font = Font(bold=True, color="1F4E78")
    ws["I13"] = "Consultants"

    bands = ["<0.3", "0.3-0.49", "0.5-0.79", ">=0.8"]
    for i, band in enumerate(bands, start=14):
        ws[f"A{i}"] = band
        ws[f"D{i}"] = band

    ws["B14"] = f"=SUMPRODUCT((Model!$A$2:$A${end_row}=1)*(Model!$B$2:$B${end_row}<0.3)*(ISNUMBER(Model!$B$2:$B${end_row})))"
    ws["B15"] = f"=SUMPRODUCT((Model!$A$2:$A${end_row}=1)*(Model!$B$2:$B${end_row}>=0.3)*(Model!$B$2:$B${end_row}<0.5))"
    ws["B16"] = f"=SUMPRODUCT((Model!$A$2:$A${end_row}=1)*(Model!$B$2:$B${end_row}>=0.5)*(Model!$B$2:$B${end_row}<0.8))"
    ws["B17"] = f"=SUMPRODUCT((Model!$A$2:$A${end_row}=1)*(ISNUMBER(Model!$B$2:$B${end_row}))*(Model!$B$2:$B${end_row}>=0.8))"

    ws["E14"] = f"=SUMPRODUCT((Model!$A$2:$A${end_row}=1)*(Model!$C$2:$C${end_row}<0.3)*(ISNUMBER(Model!$C$2:$C${end_row})))"
    ws["E15"] = f"=SUMPRODUCT((Model!$A$2:$A${end_row}=1)*(Model!$C$2:$C${end_row}>=0.3)*(Model!$C$2:$C${end_row}<0.5))"
    ws["E16"] = f"=SUMPRODUCT((Model!$A$2:$A${end_row}=1)*(Model!$C$2:$C${end_row}>=0.5)*(Model!$C$2:$C${end_row}<0.8))"
    ws["E17"] = f"=SUMPRODUCT((Model!$A$2:$A${end_row}=1)*(ISNUMBER(Model!$C$2:$C${end_row}))*(Model!$C$2:$C${end_row}>=0.8))"

    ws["H14"] = "<12m"
    ws["H15"] = "12-24m"
    ws["H16"] = ">24m"
    ws["I14"] = f"=SUMPRODUCT((Model!$A$2:$A${end_row}=1)*(Model!$D$2:$D${end_row}<>\"\")*(Model!$D$2:$D${end_row}<12))"
    ws["I15"] = f"=SUMPRODUCT((Model!$A$2:$A${end_row}=1)*(Model!$D$2:$D${end_row}>=12)*(Model!$D$2:$D${end_row}<24))"
    ws["I16"] = f"=SUMPRODUCT((Model!$A$2:$A${end_row}=1)*(Model!$D$2:$D${end_row}>=24))"

    # Segment table (Top 10)
    ws["A21"] = "Segment hotspot (Top 10 by low coverage rate)"
    ws["A21"].font = Font(bold=True, color="1F4E78")
    ws["A22"] = "Segment"
    ws["B22"] = "Low coverage rate"
    ws["C22"] = "Consultants"

    for row in range(23, 33):
        source = row - 21
        ws[f"A{row}"] = f"=Model!$O${source}"
        ws[f"B{row}"] = f"=Model!$P${source}"
        ws[f"C{row}"] = f"=Model!$Q${source}"
        ws[f"B{row}"].number_format = "0.0%"
        ws[f"X{row}"] = f"=IF(LEN(A{row})<=26,A{row},LEFT(A{row},23)&\"...\")"

    # Ranked consultant list (all selected)
    ws["A35"] = "Filtered consultant list (anomaly-first order)"
    ws["A35"].font = Font(bold=True, color="1F4E78")

    headers = [
        "Full_name",
        "Entry_date",
        "Url",
        "Department",
        "Team",
        "Lead",
        "Total_engagements",
        "Since_ratio",
        "Before_ratio",
        "Ongoing_age_m",
    ]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=36, column=idx, value=header)
        ws.cell(row=36, column=idx).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=36, column=idx).fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")

    ws["K36"] = "rank_key"

    data_start = 37
    data_end = data_start + (end_row - 2)

    for row in range(data_start, data_end + 1):
        rank = row - 36
        ws[f"K{row}"] = f"=IFERROR(LARGE(Model!$R$2:$R${end_row},{rank}),\"\")"
        ws[f"A{row}"] = f"=IF($K{row}=\"\",\"\",INDEX(Raw_data!$D$2:$D${end_row},MATCH($K{row},Model!$R$2:$R${end_row},0)))"
        ws[f"B{row}"] = f"=IF($K{row}=\"\",\"\",INDEX(Raw_data!${raw_cols['Entry date']}$2:${raw_cols['Entry date']}${end_row},MATCH($K{row},Model!$R$2:$R${end_row},0)))"
        ws[f"C{row}"] = (
            f"=IF($K{row}=\"\",\"\","
            f"HYPERLINK(INDEX(Raw_data!${raw_cols['Url']}$2:${raw_cols['Url']}${end_row},MATCH($K{row},Model!$R$2:$R${end_row},0)),"
            f"INDEX(Raw_data!${raw_cols['Url']}$2:${raw_cols['Url']}${end_row},MATCH($K{row},Model!$R$2:$R${end_row},0))))"
        )
        ws[f"D{row}"] = f"=IF($K{row}=\"\",\"\",INDEX(Raw_data!${raw_cols['Department']}$2:${raw_cols['Department']}${end_row},MATCH($K{row},Model!$R$2:$R${end_row},0)))"
        ws[f"E{row}"] = f"=IF($K{row}=\"\",\"\",INDEX(Raw_data!${raw_cols['Team']}$2:${raw_cols['Team']}${end_row},MATCH($K{row},Model!$R$2:$R${end_row},0)))"
        ws[f"F{row}"] = f"=IF($K{row}=\"\",\"\",INDEX(Raw_data!${raw_cols['Lead']}$2:${raw_cols['Lead']}${end_row},MATCH($K{row},Model!$R$2:$R${end_row},0)))"
        ws[f"G{row}"] = f"=IF($K{row}=\"\",\"\",INDEX(Raw_data!${raw_cols['Total engagements']}$2:${raw_cols['Total engagements']}${end_row},MATCH($K{row},Model!$R$2:$R${end_row},0)))"
        ws[f"H{row}"] = f"=IF($K{row}=\"\",\"\",INDEX(Model!$B$2:$B${end_row},MATCH($K{row},Model!$R$2:$R${end_row},0)))"
        ws[f"I{row}"] = f"=IF($K{row}=\"\",\"\",INDEX(Model!$C$2:$C${end_row},MATCH($K{row},Model!$R$2:$R${end_row},0)))"
        ws[f"J{row}"] = f"=IF($K{row}=\"\",\"\",INDEX(Model!$D$2:$D${end_row},MATCH($K{row},Model!$R$2:$R${end_row},0)))"
        ws[f"B{row}"].number_format = "yyyy-mm-dd"
        ws[f"H{row}"].number_format = "0.00"
        ws[f"I{row}"].number_format = "0.00"

    ws.conditional_formatting.add(
        "B14:B17",
        CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=False, fill=PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")),
    )
    ws.conditional_formatting.add(
        "E14:E17",
        CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=False, fill=PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")),
    )

    add_comments(ws)

    set_column_widths(
        ws,
        {
            "A": 30,
            "B": 14,
            "C": 26,
            "D": 32,
            "E": 30,
            "F": 22,
            "G": 16,
            "H": 12,
            "I": 12,
            "J": 12,
            "K": 3,
            "L": 4,
            "M": 18,
            "N": 18,
            "O": 18,
            "P": 18,
            "Q": 18,
            "R": 18,
            "S": 18,
            "T": 18,
            "U": 12,
            "V": 12,
            "W": 12,
            "X": 3,
            "Y": 3,
            "Z": 3,
            "AA": 3,
            "AB": 3,
            "AC": 3,
            "AD": 3,
            "AE": 3,
            "AF": 3,
            "AG": 3,
            "AH": 3,
        },
    )

    for col in ["K", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH"]:
        ws.column_dimensions[col].hidden = True

    add_dashboard_charts(ws)



def add_dashboard_charts(ws) -> None:
    band_colors = ["4F81BD", "C0504D", "9BBB59", "8064A2"]

    # 1) Since entry distribution (pie)
    since_chart = PieChart()
    since_chart.style = 2
    since_chart.title = "Since Entry Coverage Distribution"
    since_data = Reference(ws, min_col=2, min_row=13, max_row=17)
    since_categories = Reference(ws, min_col=1, min_row=14, max_row=17)
    since_chart.add_data(since_data, titles_from_data=True)
    since_chart.set_categories(since_categories)
    since_chart.legend.position = "r"
    since_chart.height = 9.2
    since_chart.width = 12.5
    if since_chart.series:
        since_chart.series[0].dLbls = DataLabelList(showVal=False, showPercent=False, showCatName=False, showSerName=False)
        for idx, color in enumerate(band_colors):
            dp = DataPoint(idx=idx)
            dp.graphicalProperties.solidFill = color
            dp.graphicalProperties.line.solidFill = color
            since_chart.series[0].data_points.append(dp)
    ws.add_chart(since_chart, "M3")

    # 2) Before entry distribution (pie)
    before_chart = PieChart()
    before_chart.style = 2
    before_chart.title = "Before Entry Coverage Distribution"
    before_data = Reference(ws, min_col=5, min_row=13, max_row=17)
    before_categories = Reference(ws, min_col=4, min_row=14, max_row=17)
    before_chart.add_data(before_data, titles_from_data=True)
    before_chart.set_categories(before_categories)
    before_chart.legend.position = "r"
    before_chart.height = 9.2
    before_chart.width = 12.5
    if before_chart.series:
        before_chart.series[0].dLbls = DataLabelList(showVal=False, showPercent=False, showCatName=False, showSerName=False)
        for idx, color in enumerate(band_colors):
            dp = DataPoint(idx=idx)
            dp.graphicalProperties.solidFill = color
            dp.graphicalProperties.line.solidFill = color
            before_chart.series[0].data_points.append(dp)
    ws.add_chart(before_chart, "M22")

    # 3) Top-10 segment low coverage (vertical bar)
    hotspot_chart = BarChart()
    hotspot_chart.type = "col"
    hotspot_chart.style = 12
    hotspot_chart.title = "Top 10 Segment Low Coverage Rate"
    hotspot_chart.y_axis.title = "Low coverage rate"
    hotspot_chart.x_axis.title = "Segment"
    hotspot_data = Reference(ws, min_col=2, min_row=22, max_row=32)
    hotspot_categories = Reference(ws, min_col=24, min_row=23, max_row=32)
    hotspot_chart.add_data(hotspot_data, titles_from_data=True)
    hotspot_chart.set_categories(hotspot_categories)
    hotspot_chart.visible_cells_only = False
    hotspot_chart.legend.position = "r"
    hotspot_chart.y_axis.scaling.min = 0
    hotspot_chart.y_axis.scaling.max = 1
    hotspot_chart.y_axis.numFmt = "0%"
    hotspot_chart.height = 11.5
    hotspot_chart.width = 16.5
    if hotspot_chart.series:
        hotspot_chart.series[0].dLbls = DataLabelList(showVal=True, showPercent=False, showCatName=False, showSerName=False)
        hotspot_chart.series[0].dLbls.numFmt = "0.0%"
        hotspot_chart.series[0].graphicalProperties.solidFill = "9BBB59"
        hotspot_chart.series[0].graphicalProperties.line.solidFill = "76923C"
    ws.add_chart(hotspot_chart, "M41")



def build_dashboard(input_path: Path, output_path: Path) -> Path:
    headers, rows = load_source(input_path)

    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    raw_cols, max_raw_row = write_raw_data_sheet(wb, headers, rows)
    list_end_rows = write_lists_sheet(wb, rows, headers)
    write_model_sheet(wb, max_raw_row, raw_cols, list_end_rows)
    write_dashboard_sheet(wb, max_raw_row, list_end_rows, raw_cols)

    wb.active = wb["Dashboard"]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path



def main() -> None:
    args = parse_args()
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser()
    if not output_path.is_absolute():
        output_path = (Path.cwd() / output_path).resolve()

    if not input_path.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")

    created = build_dashboard(input_path, output_path)
    print(f"Created dashboard workbook: {created}")


if __name__ == "__main__":
    main()
